from bs4 import BeautifulSoup
import requests
import openpyxl
excel= openpyxl.Workbook()
print(excel.sheetnames)
sheet= excel.active
sheet.title= 'top rated movies'
print(excel.sheetnames)
sheet.append(['Movie Rank', 'Movie Name', 'Year of Release', 'IMDB Rating'])

try:
    response= requests.get("https://www.imdb.com/chart/toptv/")
    response.raise_for_status()
    soup= BeautifulSoup(response.text, 'html.parser')
    tvShows= soup.find('tbody', class_='lister-list').find_all('tr')
    for tvshow in tvShows:
        rank= tvshow.find('td', class_='titleColumn').get_text(strip=True).split('.')[0]

        name= tvshow.find('td', class_='titleColumn').find('a').text
        imdbRating= tvshow.find('td', class_='ratingColumn imdbRating').find('strong').text
        YearRelease=tvshow.find('td', class_='titleColumn').find('span').text.strip('(').strip(')')
        print(rank, name,YearRelease, imdbRating)
        sheet.append([rank, name,YearRelease, imdbRating])
except Exception as e:
    print(e)
excel.save('IMDB Movei Rating.xlsx')

